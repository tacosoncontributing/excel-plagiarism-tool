from flask import Flask, request, render_template, url_for, redirect, session
from flask_bootstrap import Bootstrap
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import ARRAY
from sqlalchemy import ForeignKey
from sqlalchemy import Integer
from sqlalchemy.orm import Mapped
from sqlalchemy.orm import mapped_column
from sqlalchemy.orm import relationship
from flask_login import UserMixin, login_user, LoginManager, login_required, logout_user, current_user 
from flask_wtf import FlaskForm
from wtforms import StringField, PasswordField, SubmitField
from wtforms.validators import InputRequired, Length, ValidationError
from flask_bcrypt import Bcrypt
import os
from openpyxl import load_workbook
from datetime import datetime
from typing import List
import pythoncom
# Since this is needed to extract data on charts, the application MUST be running on Windows
import win32com.client as client

app = Flask(__name__)
Bootstrap(app)
app.config["SQLALCHEMY_DATABASE_URI"] = "sqlite:///database.db"
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SQLALCHEMY_BINDS'] = {
    'postgresql': 'postgresql://tzvupyse:uiTK_Ha5MwII2BTsuCVyIA749Ut8e4Y0@baasu.db.elephantsql.com/tzvupyse',
}
app.config["SECRET_KEY"] = "thisisasecretkey"

db = SQLAlchemy(app)
bcrypt = Bcrypt(app)

login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = "login"

class User(db.Model, UserMixin): 
  id = db.Column(db.Integer, primary_key=True)
  username = db.Column(db.String(20), nullable=False, unique=True)
  password = db.Column(db.String(80), nullable=False)

class RegistrationForm(FlaskForm): 
  username = StringField(validators=[InputRequired(), Length(min=4, max=20)], render_kw={"placeholder": "Username"})
  password = PasswordField(validators=[InputRequired(), Length(min=8, max=20)], render_kw={"placeholder": "Password"})
  submit = SubmitField("Register")

  def validate_username(self, username): 
    existing_user_username = User.query.filter_by(username=username.data).first()

    if existing_user_username: 
      raise ValidationError("This username already exists. PLease choose another one.")

class LoginForm(FlaskForm): 
  username = StringField(validators=[InputRequired(), Length(min=4, max=20)], render_kw={"placeholder": "Username"})
  password = PasswordField(validators=[InputRequired(), Length(min=8, max=20)], render_kw={"placeholder": "Password"})
  submit = SubmitField("Login")
  
# PostgreSQL database models 
class PostgreSQLUser(db.Model):
  __bind_key__ = 'postgresql'
  __tablename__ = 'postgresql_users'
  id = db.Column(db.Integer, primary_key=True)
  username = db.Column(db.String(80), unique=True, nullable=False)

class Scan(db.Model):
  __bind_key__ = 'postgresql'
  id = db.Column(db.Integer, primary_key=True)
  assignment_name = db.Column(db.String(255))
  course_name = db.Column(db.String(255))
  date_created = db.Column(db.TIMESTAMP, default=datetime.utcnow)
  number_of_files = db.Column(db.Integer)
  number_of_flagged_files = db.Column(db.Integer)
  user_created_by = db.Column(db.String(255))
  children: Mapped[List["ExcelFile"]] = relationship()
  children: Mapped[List["TemplateFile"]] = relationship()
  __tablename__ = "scans"

class ExcelFile(db.Model):
  __bind_key__ = 'postgresql'
  id = db.Column(db.Integer, primary_key=True)
  scan_id: Mapped[int] = mapped_column(ForeignKey("scans.id"))
  file_name = db.Column(db.String(255))
  created = db.Column(db.TIMESTAMP)
  creator = db.Column(db.String(255))
  modified = db.Column(db.TIMESTAMP)
  last_modified_by = db.Column(db.String(255))
  submitted_date =db.Column(db.TIMESTAMP)
  plagiarism_percentage = db.Column(db.Integer)
  unique_column_width_list = db.Column(ARRAY(db.Float))
  unique_font_names_list = db.Column(ARRAY(db.String(255)))
  complex_formulas_list =db.Column(ARRAY(db.String(255)))
  children: Mapped[List["ExcelChart"]] = relationship()
  __tablename__ = "excel_files"
  
class TemplateFile(db.Model):
  __bind_key__ = 'postgresql'
  id = db.Column(db.Integer, primary_key=True)
  scan_id: Mapped[int] = mapped_column(ForeignKey("scans.id"))
  file_name = db.Column(db.String(255))
  creator = db.Column(db.String(255))
  unique_column_width_list = db.Column(ARRAY(db.Float))
  unique_font_names_list = db.Column(ARRAY(db.String(255)))
  __tablename__ = "template_files"
  
class ExcelChart(db.Model):
  __bind_key__ = 'postgresql'
  id = db.Column(db.Integer, primary_key=True)
  excel_file_id: Mapped[int] = mapped_column(ForeignKey("excel_files.id"))
  data_source = db.Column(db.String(255))
  chart_type = db.Column(db.String(255))
  chart_left = db.Column(db.Float)
  chart_top = db.Column(db.Float)
  chart_width = db.Column(db.Float)
  chart_height = db.Column(db.Float)
  __tablename__ = "excel_charts"


# Create the databases and tables
with app.app_context():
  db.create_all()

@app.route("/login", methods=["GET", "POST"])
def login(): 
  form = LoginForm()
  if form.validate_on_submit():
    user = User.query.filter_by(username=form.username.data).first()
    if user: 
      if bcrypt.check_password_hash(user.password, form.password.data):
        login_user(user)
        return redirect(url_for("scan_list"))

  return render_template("login.html", form=form)

@app.route("/logout", methods=["GET", "POST"])
@login_required
def logout():
  logout_user()

  return redirect(url_for("login"))

@app.route("/register", methods=["GET", "POST"])
def register():
  form = RegistrationForm()

  if form.validate_on_submit():
    hashed_password = bcrypt.generate_password_hash(form.password.data)
    new_user = User(username=form.username.data, password=hashed_password)

    db.session.add(new_user)
    try:
      db.session.commit()
      print("User added successfully")  # Debugging message
    except Exception as e:
      print("Failed to add user:", e) 
      return redirect(url_for("login"))
  else:
    print("Form not validated")  # Debugging message
    print(form.errors)         
  print("Database file path:", os.path.join(os.getcwd(), "database.db"))

  return render_template("register.html", form=form)

@app.route("/scan_list")
@login_required
def scan_list(): 
  PREVIOUS_SCANS_LIST_LIMIT = 5
  previous_scans_list = Scan.query.order_by(Scan.date_created.desc()).limit(PREVIOUS_SCANS_LIST_LIMIT).all()

  return render_template("scan_list.html", previous_scans=previous_scans_list)

@app.route("/begin_scan", methods=["POST"])
@login_required
def begin_scan():
  # Assign uploaded assignment files to assignmentFiles variable
  assignment_files = request.files.getlist("assignmentFiles")
  template_file_path = get_template_file_path(request)
  
  # Initialize sets to store data from all uploaded files
  author_data = {}
  column_data = {}
  font_data = {}
  formula_data = {}
  chart_data = {}
  
  # Create a record for the newly created scan and commit that record into the PostgreSQL database
  new_scan = Scan(assignment_name=request.form.get('assignmentName'), course_name=request.form.get('courseCode'), date_created=datetime.now(), number_of_files=len(assignment_files), user_created_by=current_user.username)
  db.session.add(new_scan)
  db.session.commit()

  for file in assignment_files:
    if file:
      try:
        # Set the directory and file path where the assignment file will be saved and save it
        assignment_files_folder = "scan_assignment_uploads"
        os.makedirs(assignment_files_folder, exist_ok=True)
        assignment_file_path = os.path.join(assignment_files_folder, file.filename)
        file.save(assignment_file_path)
        
        author_data[file.filename] = get_author_data(file)
        column_data[file.filename] = get_column_data(file)
        font_data[file.filename] = get_font_names(file)
        formula_data[file.filename] = get_formula_data(file)
        chart_data[file.filename] = get_chart_data(file)

        new_file = ExcelFile(scan_id=new_scan.id,
                          file_name=file.filename,
                          created=author_data[file.filename]["created"],
                          creator=author_data[file.filename]["creator"],
                          modified=author_data[file.filename]["modified"],
                          last_modified_by=author_data[file.filename]["lastModifiedBy"],
                          submitted_date=datetime.now(),
                          plagiarism_percentage=0,
                          unique_column_width_list=column_data[file.filename],
                          unique_font_names_list=font_data[file.filename],
                          complex_formulas_list=formula_data[file.filename])
        db.session.add(new_file)
        db.session.commit()

      except Exception as e:
        return f"Error processing the file: {str(e)}"

  return redirect(url_for(".scan_results", scan_id=new_scan.id))

def get_template_file_path(request):
  template_file = None
  template_file_path = None
  # If TemplateFile is uploaded assign it to our templateFile variable to it
  if "templateFile" in request.files:
    if request.files["templateFile"].filename != "":
      template_file = request.files["templateFile"]

  # Saving the uploaded files so that they can be accessed
  if template_file is not None:
    try:
      # Set the directory and file path where the template file will be saved and save it
      template_files_folder = "scan_template_uploads"
      os.makedirs(template_files_folder, exist_ok=True)
      template_file_path = os.path.join(template_files_folder, template_file.filename)
      template_file.save(template_file_path)
    except Exception as e:
      return f"Error processing the file: {str(e)}"
  return template_file_path
  
@app.route("/scanning")
@login_required
def scanning():

  return render_template("scanning.html")

@app.route("/scan_results")
@login_required
def scan_results():
  scan_id = request.args.get("scan_id")
  scan_list = ExcelFile.query.filter_by(scan_id=scan_id).all()
  return render_template("scan_results.html", scan_list=scan_list)

@app.route("/file_details")
@login_required
def file_details():
  file_id = request.args.get('file_id')
  file = ExcelFile.query.get(file_id)
  charts = file.children
  return render_template("file_details.html", file=file, charts=charts)

@app.route("/view_scan")
@login_required
def view_scan():
  return render_template("view_scan.html")

@app.route("/settings")
@login_required
def settings():
  return render_template("settings.html")

@login_manager.user_loader
def load_user(user_id):
  return User.query.get(int(user_id))

if __name__ == "__main__":
  app.run(host="127.0.0.1", Pport=8080, debug=True)

def get_column_data(excel_file):
  file_column_data = set()
  # Go through each file in the given list of excel files
  try:
    # If the file has no filename, something went wrong
    if excel_file.filename == "":
      print(f"Could not retrieve filename from {excel_file}")
    # Otherwise save the file, open the workbook, go through each sheet and store the unique column widths (in each respective sheet)
    else:
      if excel_file:
        assignment_files_folder = "scan_assignment_uploads"
        assignment_file_path = os.path.join(assignment_files_folder, excel_file.filename)
        excel_workbook = load_workbook(assignment_file_path)
        for sheet_name in excel_workbook.sheetnames:
          excel_sheet = excel_workbook[sheet_name]
          for column in excel_sheet.columns:
            width = excel_sheet.column_dimensions[column[0].column_letter].width
            file_column_data.add(width)
        excel_workbook.close()
  except Exception as e:
    print(f"Error reading {excel_file}: {str(e)}")

  return list(file_column_data)

def get_author_data(excel_file):
  file_author_data = {}
  try:
    # If the file has no filename, something went wrong
    if excel_file.filename == "":
      print(f"Could not retrieve filename from {excel_file}")
    else:
      # Otherwise save the file, open the workbook, and store the workbook properties (contains file metadata like creator, title, description, createdDate, lastModifiedDate, lastModifiedBy, etc)
      if excel_file: 
        assignment_files_folder = "scan_assignment_uploads"
        assignment_file_path = os.path.join(assignment_files_folder, excel_file.filename)
        excel_workbook = load_workbook(assignment_file_path)
        file_author_data["creator"] = excel_workbook.properties.creator
        file_author_data["created"] = excel_workbook.properties.created
        file_author_data["modified"] = excel_workbook.properties.modified
        file_author_data["lastModifiedBy"] = excel_workbook.properties.lastModifiedBy
        excel_workbook.close()
  except Exception as e:
    print(f"Error reading {excel_file}: {str(e)}")

  return file_author_data

def get_font_names(excel_file):
  font_names_data = []
  try:
    # If the file has no filename, something went wrong
    if excel_file.filename == "":
      print(f"Could not retrieve filename from {excel_file}")
    else:
      # Otherwise save the file, open the workbook, and get every unique font from each file and store them into a list (which contains all the unique fonts used by each excel file)
      if excel_file:
        assignment_files_folder = "scan_assignment_uploads"
        assignment_file_path = os.path.join(assignment_files_folder, excel_file.filename)
        excel_workbook = load_workbook(assignment_file_path)
        for sheet_name in excel_workbook.sheetnames:
          excel_sheet = excel_workbook[sheet_name]
          for row in excel_sheet.iter_rows(min_row=1, max_col=excel_sheet.max_column, max_row=excel_sheet.max_row):
            for cell in row:
              font = cell.font
              if font.name not in font_names_data: 
                font_names_data.append(font.name)
        excel_workbook.close()
  except Exception as e:
    print(f"Error reading {excel_file}: {str(e)}")

  return font_names_data

def get_chart_data(excel_file):
  file_chart_data = []
  # Go through each file in the given list of excel files
  try:
    # If the file has no filename, something went wrong
    if excel_file.filename == "":
      print(f"Could not retrieve filename from {excel_file}")
    else:
      pythoncom.CoInitialize()
      assignment_file_path = get_absolute_path(excel_file.filename)
      excel_app = client.Dispatch('Excel.Application')
      excel_workbook = excel_app.Workbooks.Open(assignment_file_path)
      for sheet in excel_workbook.Sheets:
        # Code to indicate that sheet is a CHART SHEET (contains only charts)
        if sheet.Type == -4100:
          series_output(sheet)
        # otherwise it's a regular worksheet (which may contain charts)
        elif sheet.Type == -4167:
          for chart in sheet.ChartObjects():
            file_chart_data.append(series_output(chart.Chart))
      # Don't save and close workbook (otherwise charts will automatically be removed)
      excel_workbook.Close()
      excel_app.Quit()
  except Exception as e:
    print(f"Error reading {excel_file}: {str(e)}")

  return file_chart_data

def series_output(chart):
  chart_data = {
      "Chart Name": chart.Name,
      "Chart Type": chart.ChartType,
      "Series": []
    }
  for series in chart.SeriesCollection():
    chart_data["Series"].append({
        "Name": series.Name,
        "Formula": series.Formula
      })
  return chart_data
    
def get_absolute_path(filename):
  # Get the current directory of the script
  current_dir = os.path.dirname(os.path.abspath(__file__))
  # Construct the absolute path using the current directory and the filename
  absolute_path = os.path.join(current_dir, "scan_assignment_uploads", filename)

  return absolute_path

def get_formula_data(excel_file):
  file_formula_data = []
  try:
    # If the file has no filename, something went wrong
    if excel_file.filename == "":
      print(f"Could not retrieve filename from {excel_file}")
    else:
      # Otherwise save the file, open the workbook, and get the formula from every cell which contains a formula
      if excel_file:
        assignment_files_folder = "scan_assignment_uploads"
        assignment_file_path = os.path.join(assignment_files_folder, excel_file.filename)
        excel_workbook = load_workbook(assignment_file_path)
        for sheet_name in excel_workbook.sheetnames:
          excel_sheet = excel_workbook[sheet_name]
          for row in excel_sheet.iter_rows(min_row=1, max_col=excel_sheet.max_column, max_row=excel_sheet.max_row):
            for cell in row:
              if cell.data_type == "f":
                file_formula_data.append(cell.value)
  except Exception as e:
    print(f"Error reading {file_formula_data}: {str(e)}")

  return file_formula_data
